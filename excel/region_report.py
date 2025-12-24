# ==================== REGION REPORT ====================
# Bölge özet Excel raporu

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from .excel_utils import (
    auto_adjust_column_width, HEADER_FONT, HEADER_FILL, 
    TITLE_FONT, BORDER, get_risk_fill,
    KRITIK_FILL, RISKLI_FILL, DIKKAT_FILL, TEMIZ_FILL
)


def create_region_excel_report(region_df, df_all, kasa_kodlari, params):
    """Bölge özet Excel raporu"""
    
    wb = Workbook()
    
    # ===== BÖLGE ÖZETİ =====
    ws = wb.active
    ws.title = "BÖLGE ÖZETİ"
    
    ws['A1'] = f"BÖLGE ENVANTER ANALİZİ"
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"Dönem: {params.get('donem', '')} | Tarih: {params.get('tarih', '')} | Mağaza Sayısı: {len(region_df)}"
    
    # Bölge toplamları
    ws['A4'] = "BÖLGE TOPLAMI"
    ws['A4'].font = Font(bold=True, size=11)
    
    toplam_satis = region_df['Satış'].sum()
    toplam_fark = region_df['Fark'].sum()
    toplam_fire = region_df['Fire'].sum()
    genel_oran = abs(toplam_fark + toplam_fire) / toplam_satis * 100 if toplam_satis > 0 else 0
    
    ws['A5'] = "Toplam Satış"
    ws['B5'] = f"{toplam_satis:,.0f} TL"
    ws['A6'] = "Toplam Fark"
    ws['B6'] = f"{toplam_fark:,.0f} TL"
    ws['A7'] = "Toplam Fire"
    ws['B7'] = f"{toplam_fire:,.0f} TL"
    ws['A8'] = "Genel Kayıp Oranı"
    ws['B8'] = f"%{genel_oran:.2f}"
    
    # Risk dağılımı
    ws['A10'] = "RİSK DAĞILIMI"
    ws['A10'].font = Font(bold=True, size=11)
    
    kritik_sayisi = len(region_df[region_df['Risk'].str.contains('KRİTİK', na=False)])
    riskli_sayisi = len(region_df[region_df['Risk'].str.contains('RİSKLİ', na=False)])
    dikkat_sayisi = len(region_df[region_df['Risk'].str.contains('DİKKAT', na=False)])
    temiz_sayisi = len(region_df[region_df['Risk'].str.contains('TEMİZ', na=False)])
    
    ws['A11'] = "🔴 KRİTİK"
    ws['B11'] = kritik_sayisi
    ws['A12'] = "🟠 RİSKLİ"
    ws['B12'] = riskli_sayisi
    ws['A13'] = "🟡 DİKKAT"
    ws['B13'] = dikkat_sayisi
    ws['A14'] = "🟢 TEMİZ"
    ws['B14'] = temiz_sayisi
    
    # Mağaza sıralaması
    ws['A16'] = "MAĞAZA SIRALAMASI (Risk Puanına Göre)"
    ws['A16'].font = Font(bold=True, size=11)
    
    headers = ['Mağaza', 'Adı', 'Satış', 'Fark', 'Toplam %', 'İç Hırs.', 'Sigara', 'Kr.Açık', 'Risk', 'Neden']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=17, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    
    for row_idx, (_, row) in enumerate(region_df.iterrows(), start=18):
        ws.cell(row=row_idx, column=1, value=row['Mağaza Kodu']).border = BORDER
        ws.cell(row=row_idx, column=2, value=str(row['Mağaza Adı'])[:25]).border = BORDER
        ws.cell(row=row_idx, column=3, value=f"{row['Satış']:,.0f}").border = BORDER
        ws.cell(row=row_idx, column=4, value=f"{row['Fark']:,.0f}").border = BORDER
        ws.cell(row=row_idx, column=5, value=f"%{row['Toplam %']:.1f}").border = BORDER
        ws.cell(row=row_idx, column=6, value=row['İç Hırs.']).border = BORDER
        ws.cell(row=row_idx, column=7, value=row['Sigara']).border = BORDER
        ws.cell(row=row_idx, column=8, value=row['Kr.Açık']).border = BORDER
        
        risk_cell = ws.cell(row=row_idx, column=9, value=row['Risk'])
        risk_cell.border = BORDER
        risk_cell.fill = get_risk_fill(row['Risk'])
        risk_cell.font = Font(bold=True, color='FFFFFF')
        
        ws.cell(row=row_idx, column=10, value=row.get('Risk Nedenleri', '')).border = BORDER
    
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 35
    
    # ===== DETAY SHEET =====
    ws2 = wb.create_sheet("DETAY")
    
    detail_headers = ['Mağaza Kodu', 'Mağaza Adı', 'Satış', 'Fark', 'Fire', 'Toplam %', 
                      'İç Hırs.', 'Kr.Açık', 'Kr.Fire', 'Sigara', 'Fire Man.', 
                      '10TL Adet', '10TL Tutar', 'Risk Puan', 'Risk', 'Risk Nedenleri']
    
    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    
    for row_idx, (_, row) in enumerate(region_df.iterrows(), start=2):
        ws2.cell(row=row_idx, column=1, value=row['Mağaza Kodu']).border = BORDER
        ws2.cell(row=row_idx, column=2, value=row['Mağaza Adı']).border = BORDER
        ws2.cell(row=row_idx, column=3, value=row['Satış']).border = BORDER
        ws2.cell(row=row_idx, column=4, value=row['Fark']).border = BORDER
        ws2.cell(row=row_idx, column=5, value=row['Fire']).border = BORDER
        ws2.cell(row=row_idx, column=6, value=row['Toplam %']).border = BORDER
        ws2.cell(row=row_idx, column=7, value=row['İç Hırs.']).border = BORDER
        ws2.cell(row=row_idx, column=8, value=row['Kr.Açık']).border = BORDER
        ws2.cell(row=row_idx, column=9, value=row.get('Kr.Fire', 0)).border = BORDER
        ws2.cell(row=row_idx, column=10, value=row['Sigara']).border = BORDER
        ws2.cell(row=row_idx, column=11, value=row.get('Fire Man.', 0)).border = BORDER
        ws2.cell(row=row_idx, column=12, value=row['10TL Adet']).border = BORDER
        ws2.cell(row=row_idx, column=13, value=row['10TL Tutar']).border = BORDER
        ws2.cell(row=row_idx, column=14, value=row['Risk Puan']).border = BORDER
        ws2.cell(row=row_idx, column=15, value=row['Risk']).border = BORDER
        ws2.cell(row=row_idx, column=16, value=row.get('Risk Nedenleri', '')).border = BORDER
    
    auto_adjust_column_width(ws2)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
