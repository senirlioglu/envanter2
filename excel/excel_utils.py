# ==================== EXCEL UTILS ====================
# Excel yardımcı fonksiyonlar

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Stiller
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
KRITIK_FILL = PatternFill('solid', fgColor='FF4444')
RISKLI_FILL = PatternFill('solid', fgColor='FF8800')
DIKKAT_FILL = PatternFill('solid', fgColor='FFCC00')
TEMIZ_FILL = PatternFill('solid', fgColor='00CC66')
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11)
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical='top')


def auto_adjust_column_width(ws):
    """Excel sütun genişliklerini otomatik ayarla"""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width


def get_risk_fill(risk_text):
    """Risk seviyesine göre renk döndür"""
    risk_str = str(risk_text)
    if 'KRİTİK' in risk_str:
        return KRITIK_FILL
    elif 'RİSKLİ' in risk_str:
        return RISKLI_FILL
    elif 'DİKKAT' in risk_str:
        return DIKKAT_FILL
    else:
        return TEMIZ_FILL
