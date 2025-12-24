# ==================== GM REPORT ====================
# GM Dashboard Excel raporu

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
from .excel_utils import (
    auto_adjust_column_width, HEADER_FONT, HEADER_FILL, 
    TITLE_FONT, BORDER, get_risk_fill
)


def create_gm_excel_report(store_df, sm_df, bs_df, params):
    """GM Dashboard Excel raporu"""
    
    wb = Workbook()
    
    # ===== BÖLGE ÖZETİ =====
    ws = wb.active
    ws.title = "BÖLGE ÖZETİ"
    
    ws['A1'] = "GM BÖLGE DASHBOARD"
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"Dönem: {params.get('donem', '')} | Mağaza: {len(store_df)}"
    
    # Toplamlar
    toplam_satis = store_df['Satış'].sum()
    toplam_fark = store_df['Fark'].sum()
    toplam_fire = store_df['Fire'].sum()
    toplam_acik = store_df['Toplam Açık'].sum()
    
    ws['A4'] = "GENEL METRİKLER"
    ws['A4'].font = Font(bold=True, size=11)
    
    ws['A5'] = "Toplam Satış"
    ws['B5'] = f"{toplam_satis:,.0f} TL"
    ws['A6'] = "Toplam Fark"
    ws['B6'] = f"{toplam_fark:,.0f} TL"
    ws['A7'] = "Toplam Fire"
    ws['B7'] = f"{toplam_fire:,.0f} TL"
    ws['A8'] = "Toplam Açık"
    ws['B8'] = f"{toplam_acik:,.0f} TL"
    ws['A9'] = "Kayıp Oranı"
    ws['B9'] = f"%{abs(toplam_acik)/toplam_satis*100:.2f}" if toplam_satis > 0 else "0%"
    
    # Risk dağılımı
    ws['A11'] = "RİSK DAĞILIMI"
    ws['A11'].font = Font(bold=True, size=11)
    
    kritik = len(store_df[store_df['Risk'].str.contains('KRİTİK', na=False)])
    riskli = len(store_df[store_df['Risk'].str.contains('RİSKLİ', na=False)])
    dikkat = len(store_df[store_df['Risk'].str.contains('DİKKAT', na=False)])
    temiz = len(store_df[store_df['Risk'].str.contains('TEMİZ', na=False)])
    
    ws['A12'] = "🔴 KRİTİK"
    ws['B12'] = kritik
    ws['A13'] = "🟠 RİSKLİ"
    ws['B13'] = riskli
    ws['A14'] = "🟡 DİKKAT"
    ws['B14'] = dikkat
    ws['A15'] = "🟢 TEMİZ"
    ws['B15'] = temiz
    
    # ===== SM BAZLI =====
    if len(sm_df) > 0:
        ws2 = wb.create_sheet("SM BAZLI")
        headers = ['Satış Müdürü', 'Mağaza', 'Satış', 'Fark', 'Fire', 'Toplam %', 'Sigara', 'İç Hırs.', 'Ort. Risk']
        
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
        
        for row_idx, (_, row) in enumerate(sm_df.iterrows(), start=2):
            ws2.cell(row=row_idx, column=1, value=row.get('SM', '')).border = BORDER
            ws2.cell(row=row_idx, column=2, value=row.get('Mağaza Sayısı', 0)).border = BORDER
            ws2.cell(row=row_idx, column=3, value=f"{row['Satış']:,.0f}").border = BORDER
            ws2.cell(row=row_idx, column=4, value=f"{row['Fark']:,.0f}").border = BORDER
            ws2.cell(row=row_idx, column=5, value=f"{row['Fire']:,.0f}").border = BORDER
            ws2.cell(row=row_idx, column=6, value=f"%{row.get('Toplam %', 0):.1f}").border = BORDER
            ws2.cell(row=row_idx, column=7, value=row.get('Sigara', 0)).border = BORDER
            ws2.cell(row=row_idx, column=8, value=row.get('İç Hırs.', 0)).border = BORDER
            ws2.cell(row=row_idx, column=9, value=f"{row.get('Ort. Risk', 0):.0f}").border = BORDER
        
        auto_adjust_column_width(ws2)
    
    # ===== BS BAZLI =====
    if len(bs_df) > 0:
        ws3 = wb.create_sheet("BS BAZLI")
        headers = ['Bölge Sorumlusu', 'Mağaza', 'Satış', 'Fark', 'Fire', 'Toplam %', 'Sigara', 'İç Hırs.', 'Ort. Risk']
        
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
        
        for row_idx, (_, row) in enumerate(bs_df.iterrows(), start=2):
            ws3.cell(row=row_idx, column=1, value=row.get('BS', '')).border = BORDER
            ws3.cell(row=row_idx, column=2, value=row.get('Mağaza Sayısı', 0)).border = BORDER
            ws3.cell(row=row_idx, column=3, value=f"{row['Satış']:,.0f}").border = BORDER
            ws3.cell(row=row_idx, column=4, value=f"{row['Fark']:,.0f}").border = BORDER
            ws3.cell(row=row_idx, column=5, value=f"{row['Fire']:,.0f}").border = BORDER
            ws3.cell(row=row_idx, column=6, value=f"%{row.get('Toplam %', 0):.1f}").border = BORDER
            ws3.cell(row=row_idx, column=7, value=row.get('Sigara', 0)).border = BORDER
            ws3.cell(row=row_idx, column=8, value=row.get('İç Hırs.', 0)).border = BORDER
            ws3.cell(row=row_idx, column=9, value=f"{row.get('Ort. Risk', 0):.0f}").border = BORDER
        
        auto_adjust_column_width(ws3)
    
    # ===== TÜM MAĞAZALAR =====
    ws4 = wb.create_sheet("TÜM MAĞAZALAR")
    headers = ['Mağaza Kodu', 'Mağaza Adı', 'SM', 'BS', 'Satış', 'Fark', 'Fire', 'Toplam %', 
               'Sigara', 'İç Hırs.', '10TL Adet', 'Risk Puan', 'Risk', 'Nedenler']
    
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    
    for row_idx, (_, row) in enumerate(store_df.iterrows(), start=2):
        ws4.cell(row=row_idx, column=1, value=row['Mağaza Kodu']).border = BORDER
        ws4.cell(row=row_idx, column=2, value=row['Mağaza Adı']).border = BORDER
        ws4.cell(row=row_idx, column=3, value=row.get('SM', '')).border = BORDER
        ws4.cell(row=row_idx, column=4, value=row.get('BS', '')).border = BORDER
        ws4.cell(row=row_idx, column=5, value=f"{row['Satış']:,.0f}").border = BORDER
        ws4.cell(row=row_idx, column=6, value=f"{row['Fark']:,.0f}").border = BORDER
        ws4.cell(row=row_idx, column=7, value=f"{row['Fire']:,.0f}").border = BORDER
        ws4.cell(row=row_idx, column=8, value=f"%{row.get('Toplam %', 0):.1f}").border = BORDER
        ws4.cell(row=row_idx, column=9, value=row.get('Sigara', 0)).border = BORDER
        ws4.cell(row=row_idx, column=10, value=row.get('İç Hırs.', 0)).border = BORDER
        
        kasa_adet = row.get('Kasa Adet', row.get('10TL Adet', 0))
        ws4.cell(row=row_idx, column=11, value=kasa_adet).border = BORDER
        ws4.cell(row=row_idx, column=12, value=f"{row.get('Risk Puan', 0):.0f}").border = BORDER
        
        risk_cell = ws4.cell(row=row_idx, column=13, value=row['Risk'])
        risk_cell.border = BORDER
        risk_cell.fill = get_risk_fill(row['Risk'])
        
        ws4.cell(row=row_idx, column=14, value=row.get('Risk Nedenleri', '')).border = BORDER
    
    auto_adjust_column_width(ws4)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
