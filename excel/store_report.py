# ==================== STORE REPORT ====================
# Tek mağaza Excel raporu

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from .excel_utils import (
    auto_adjust_column_width, HEADER_FONT, HEADER_FILL, 
    TITLE_FONT, SUBTITLE_FONT, BORDER, WRAP_ALIGNMENT
)


def create_excel_report(df, internal_df, chronic_df, chronic_fire_df, cigarette_df, 
                       external_df, family_df, fire_manip_df, kasa_activity_df, top20_df, 
                       exec_comments, group_stats, magaza_kodu, magaza_adi, params):
    """Excel raporu - tüm sheet'ler dahil"""
    
    wb = Workbook()
    
    # ===== ÖZET =====
    ws = wb.active
    ws.title = "ÖZET"
    
    ws['A1'] = f"MAĞAZA: {magaza_kodu} - {magaza_adi}"
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"Dönem: {params.get('donem', '')} | Tarih: {params.get('tarih', '')}"
    
    ws['A4'] = "GENEL METRİKLER"
    ws['A4'].font = SUBTITLE_FONT
    
    toplam_satis = df['Satış Tutarı'].sum() if 'Satış Tutarı' in df.columns else 0
    fark_tutari = df['Fark Tutarı'].fillna(0).sum() if 'Fark Tutarı' in df.columns else 0
    kismi_tutari = df['Kısmi Envanter Tutarı'].fillna(0).sum() if 'Kısmi Envanter Tutarı' in df.columns else 0
    fire_tutari = df['Fire Tutarı'].fillna(0).sum() if 'Fire Tutarı' in df.columns else 0
    
    fark = fark_tutari + kismi_tutari
    toplam_acik = fark + fire_tutari
    
    fark_oran = abs(fark) / toplam_satis * 100 if toplam_satis > 0 else 0
    fire_oran = abs(fire_tutari) / toplam_satis * 100 if toplam_satis > 0 else 0
    toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
    
    metrics = [
        ('Toplam Ürün', len(df)),
        ('Açık Veren Ürün', len(df[df['Fark Miktarı'] < 0]) if 'Fark Miktarı' in df.columns else 0),
        ('Toplam Satış', f"{toplam_satis:,.0f} TL"),
        ('Fark (Fark+Kısmi)', f"{fark:,.0f} TL"),
        ('Fire', f"{fire_tutari:,.0f} TL"),
        ('Toplam Açık', f"{toplam_acik:,.0f} TL"),
        ('Fark Oranı', f"%{fark_oran:.2f}"),
        ('Fire Oranı', f"%{fire_oran:.2f}"),
        ('Toplam Oran', f"%{toplam_oran:.2f}"),
    ]
    
    for i, (label, value) in enumerate(metrics, start=5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
    
    ws['A15'] = "RİSK DAĞILIMI"
    ws['A15'].font = SUBTITLE_FONT
    
    sigara_net_toplam = 0
    if len(cigarette_df) > 0 and 'Ürün Toplam' in cigarette_df.columns:
        toplam_row = cigarette_df[cigarette_df['Malzeme Kodu'] == '*** TOPLAM ***']
        if len(toplam_row) > 0:
            sigara_net_toplam = abs(toplam_row['Ürün Toplam'].values[0])
    
    risks = [
        ('İç Hırsızlık (≥100TL)', len(internal_df)),
        ('Kronik Açık', len(chronic_df)),
        ('Kronik Fire', len(chronic_fire_df)),
        ('Sigara Açığı', int(sigara_net_toplam)),
        ('Fire Manipülasyonu', len(fire_manip_df)),
    ]
    
    for i, (label, value) in enumerate(risks, start=16):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        if 'Sigara' in label and value > 0:
            ws[f'B{i}'].fill = PatternFill('solid', fgColor='FF4444')
            ws[f'B{i}'].font = Font(bold=True, color='FFFFFF')
    
    ws['A22'] = "YÖNETİCİ ÖZETİ"
    ws['A22'].font = SUBTITLE_FONT
    
    for i, comment in enumerate(exec_comments[:10], start=23):
        ws[f'A{i}'] = comment
    
    auto_adjust_column_width(ws)
    
    # ===== EN RİSKLİ 20 =====
    if len(top20_df) > 0:
        ws2 = wb.create_sheet("EN RİSKLİ 20")
        for col, h in enumerate(top20_df.columns, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
        
        for r_idx, row in enumerate(top20_df.values, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                cell.border = BORDER
                cell.alignment = WRAP_ALIGNMENT
        
        auto_adjust_column_width(ws2)
    
    # ===== KRONİK AÇIK =====
    if len(chronic_df) > 0:
        ws3 = wb.create_sheet("KRONİK AÇIK")
        for col, h in enumerate(chronic_df.columns, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        
        for r_idx, row in enumerate(chronic_df.head(100).values, 2):
            for c_idx, val in enumerate(row, 1):
                ws3.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws3)
    
    # ===== KRONİK FİRE =====
    if len(chronic_fire_df) > 0:
        ws4 = wb.create_sheet("KRONİK FİRE")
        for col, h in enumerate(chronic_fire_df.columns, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        
        for r_idx, row in enumerate(chronic_fire_df.head(100).values, 2):
            for c_idx, val in enumerate(row, 1):
                ws4.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws4)
    
    # ===== SİGARA AÇIĞI =====
    ws5 = wb.create_sheet("SİGARA AÇIĞI")
    ws5['A1'] = "⚠️ SİGARA AÇIĞI - YÜKSEK RİSK"
    ws5['A1'].font = Font(bold=True, size=14, color='FF0000')
    
    if len(cigarette_df) > 0:
        for col, h in enumerate(cigarette_df.columns, 1):
            cell = ws5.cell(row=3, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = PatternFill('solid', fgColor='FF4444')
        
        for r_idx, row in enumerate(cigarette_df.values, 4):
            for c_idx, val in enumerate(row, 1):
                ws5.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws5)
    
    # ===== İÇ HIRSIZLIK =====
    if len(internal_df) > 0:
        ws6 = wb.create_sheet("İÇ HIRSIZLIK")
        ws6['A1'] = "Satış Fiyatı ≥ 100 TL | Fark büyüdükçe risk AZALIR"
        ws6['A1'].font = SUBTITLE_FONT
        
        for col, h in enumerate(internal_df.columns, 1):
            cell = ws6.cell(row=3, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        
        for r_idx, row in enumerate(internal_df.head(100).values, 4):
            for c_idx, val in enumerate(row, 1):
                ws6.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws6)
    
    # ===== AİLE ANALİZİ =====
    if len(family_df) > 0:
        ws7 = wb.create_sheet("AİLE ANALİZİ")
        ws7['A1'] = "Benzer Ürün Ailesi - Kod Karışıklığı Tespiti"
        ws7['A1'].font = SUBTITLE_FONT
        
        for col, h in enumerate(family_df.columns, 1):
            cell = ws7.cell(row=3, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        
        for r_idx, row in enumerate(family_df.head(100).values, 4):
            for c_idx, val in enumerate(row, 1):
                cell = ws7.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = WRAP_ALIGNMENT
        
        auto_adjust_column_width(ws7)
    
    # ===== FİRE MANİPÜLASYONU =====
    if len(fire_manip_df) > 0:
        ws8 = wb.create_sheet("FİRE MANİPÜLASYONU")
        for col, h in enumerate(fire_manip_df.columns, 1):
            cell = ws8.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        
        for r_idx, row in enumerate(fire_manip_df.head(100).values, 2):
            for c_idx, val in enumerate(row, 1):
                ws8.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws8)
    
    # ===== KASA AKTİVİTESİ =====
    if len(kasa_activity_df) > 0:
        ws9 = wb.create_sheet("KASA AKTİVİTESİ")
        ws9['A1'] = "⚠️ KASA AKTİVİTESİ ÜRÜNLERİ"
        ws9['A1'].font = Font(bold=True, size=12, color='FF0000')
        
        for col, h in enumerate(kasa_activity_df.columns, 1):
            cell = ws9.cell(row=3, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        
        for r_idx, row in enumerate(kasa_activity_df.values, 4):
            for c_idx, val in enumerate(row, 1):
                cell = ws9.cell(row=r_idx, column=c_idx, value=val)
                if c_idx == 5 and isinstance(val, (int, float)) and val > 0:
                    cell.fill = PatternFill('solid', fgColor='FFCCCC')
        
        auto_adjust_column_width(ws9)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
